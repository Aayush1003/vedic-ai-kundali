import openpyxl

wb = openpyxl.load_workbook('docs/Kundali Milan (1).xlsx')
print('Sheet names:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    print()
    for row in ws.iter_rows(min_row=1, values_only=False):
        for cell in row:
            print(f'  {cell.coordinate}: {repr(cell.value)}')
        print('---')
